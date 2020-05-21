import pandas as pd
import xlrd
import datetime
import xlwt
import openpyxl
#D:\19ПИ-3.xlsx - это путь к файлу
class STUDENT(object):
    def __init__(self, name='', group='', points=[], midpoint=int, respoint=int):
        self.name = name
        self.group = group
        self.points = points
        self.midpoint = midpoint
        self.respoint = respoint

    def show_student(self):
        discription = (str(self.name) + " " + str(self.group) + " middle point is " + str(
            self.midpoint) + " result point is " + str(self.respoint))
        print(discription)
def Сurrent_student_grades(file_name : str):
    if type(file_name) is str:
        students = []
        print('Enter filename(number of group): ')
        filename = str(input())
        group = filename
        rb = xlrd.open_workbook(file_name, formatting_info=False)
        if rb == 0:
            print('Open error')
        sheet = rb.sheet_by_index(0)            #выбираем активный лист
        row = []
        row_number = sheet.nrows
        col_number = sheet.ncols
        i = 2
        kolvo = 0
        while sheet.cell(rowx=0, colx=i).value != 'Средний балл' and sheet.cell(rowx=1, colx=i).value != 'Middle point':
            kolvo += 1
            i += 1
        k = 0
        if row_number > 0:
            for i in range(1, row_number):
                array = [0]*kolvo
                j = 1
                k = 0
                while sheet.cell(rowx=0, colx=j+1).value != 'Средний балл' and sheet.cell(rowx=1, colx=j).value != 'Middle point':
                    j += 1
                    array[k] = str(sheet.cell(rowx=i, colx=j).value)
                    k+=1
                Student = STUDENT(sheet.cell(rowx=i, colx=1).value, group, array, sheet.cell(rowx=i, colx=kolvo+2).value,
                                  sheet.cell(rowx=i, colx=1+kolvo+2).value)
                students.append(Student)
                #Student.show_student()
                array.clear()
        else:
                return -1
        return students
    else:
        return -1
print('Enter the way to your file(with name of this file): ')
way_ = str(input())
arr = []
arr = Сurrent_student_grades(way_)
i = 0
for i in range (len(arr)):
    arr[i].show_student()